from django.shortcuts import render
from django.views.generic import TemplateView, CreateView, UpdateView, DeleteView, ListView, DetailView, View
from django.urls import reverse_lazy
from django.http import HttpResponseRedirect
from django.shortcuts import render, redirect
from aplicaciones.ctr_escolar.models import Aula, Materia, Documento, Unidad, Tarea, Blog, \
TareaDocumento, CalificacionUnidad, CalificacionMateria, ComentarioBlog
from aplicaciones.ctr_escolar.forms import AulaForm, MateriaForm,MateriaFormEdit, DocumentoCreateForm, UnidadForm, \
TareaForm, TareaFormEdit, BlogFrom, TareaDocumentoFrom, TareaEntregadaEdit, ComentarioBlogForm
from aplicaciones.ctr_escolar.eliminaciones import get_deleted_objects
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import IntegrityError
from django.shortcuts import render_to_response
from datetime import datetime, date, timedelta
from django.db.models import Sum, FloatField, F
import pytz
from django.utils import timezone 

from django.core import serializers

from django.core.serializers.json import DjangoJSONEncoder
from django.core.serializers import serialize

from openpyxl.styles import Font, Fill, Alignment
from django.http import HttpResponse
from openpyxl import Workbook
"""
CLASES ESCUSIVAMENTE PARA LOS USUARIOS CON PERMISOS DE MAESTRO
DE AQUI EN ADELANTE
#########################################################################################################
#########################################################################################################
#########################################################################################################
"""
from django.http import JsonResponse
# Create your views here.
class index(LoginRequiredMixin, TemplateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    template_name = 'admin/base_sensei.html'

    def get_context_data(self, **kwargs):
        from datetime import datetime,timedelta
        hoy = datetime.now()#fecha actual
        dias = timedelta(days=10)
        hoy_mas_10_dias = hoy+dias
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user 
        if self.request.user.is_maestro:
            context['proximas_tareas'] = Tarea.objects.filter(tarea_unidad__unidad_materia__materia_aula__aula_pertenece=self.request.user, tarea_fecha_termino__range=(hoy, hoy_mas_10_dias)).order_by('tarea_fecha_termino')[:10]
            
            context['total_tareas'] = Tarea.objects.filter(tarea_unidad__unidad_materia__materia_aula__aula_pertenece=self.request.user).count()
            context['tareas_pasadas'] = Tarea.objects.filter(tarea_unidad__unidad_materia__materia_aula__aula_pertenece=self.request.user, tarea_fecha_termino__lt=hoy).count()
           
            context['tareas_entragadas_total'] = TareaDocumento.objects.filter(tareaDocumento_Tarea__tarea_unidad__unidad_materia__materia_aula__aula_pertenece=self.request.user).count()
            context['tareas_calificadas'] = TareaDocumento.objects.filter(tareaDocumento_Tarea__tarea_unidad__unidad_materia__materia_aula__aula_pertenece=self.request.user, tareaDocumento_status=True).count()
        else:
            context['proximas_tareas'] = Tarea.objects.filter(tarea_unidad__unidad_materia__materia_registro_alumnnos=self.request.user, tarea_fecha_termino__range=(hoy, hoy_mas_10_dias)).order_by('tarea_fecha_termino')[:10]
            
            context['total_tareas'] = Tarea.objects.filter(tarea_unidad__unidad_materia__materia_registro_alumnnos=self.request.user).count()
            context['tareas_pasadas'] = Tarea.objects.filter(tarea_unidad__unidad_materia__materia_registro_alumnnos=self.request.user, tarea_fecha_termino__lt=hoy).count()
           
            context['tareas_entragadas_total'] = TareaDocumento.objects.filter(tareaDocumento_pertenece=self.request.user).count()
            context['tareas_calificadas'] = TareaDocumento.objects.filter(tareaDocumento_pertenece=self.request.user, tareaDocumento_status=True).count()

        return context


class CreateAula(LoginRequiredMixin, CreateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Aula
    form_class = AulaForm
    template_name = 'maestro/create_aula.html'
    success_url = reverse_lazy('control_escolar:maestro_create_aula')

    def form_valid(self, form):
        form.instance.aula_pertenece = self.request.user
        return super(CreateAula, self).form_valid(form)

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'aula'
        context['list_aula'] = Aula.objects.filter(aula_pertenece=self.request.user)
        return context


class UpdateAula(LoginRequiredMixin, UpdateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Aula
    form_class = AulaForm
    template_name = 'maestro/create_aula.html'
    success_url = reverse_lazy('control_escolar:maestro_create_aula')

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'aula'
        context['list_aula'] = Aula.objects.filter(aula_pertenece=self.request.user)

        return context

class AulaDelete(LoginRequiredMixin, DeleteView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'

    model = Aula
    template_name = 'eliminaciones.html'
    success_url = reverse_lazy('control_escolar:maestro_create_aula')

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'aula'
        context['list_aula'] = Aula.objects.filter(aula_pertenece=self.request.user)

        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected

        return context

class MateriaCreate(LoginRequiredMixin, CreateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Materia
    form_class = MateriaForm
    template_name = 'maestro/create_materia.html'
    success_url = reverse_lazy('control_escolar:maestro_materias')

    def get_form_kwargs(self):
        kwargs = super(MateriaCreate, self).get_form_kwargs()
        kwargs.update({'user': self.request.user}) 
        return kwargs

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'materia'
        context['lista'] = Materia.objects.filter(materia_aula__aula_pertenece=self.request.user)

        return context

class MateriaUpdate(LoginRequiredMixin, UpdateView):
    model = Materia
    form_class = MateriaFormEdit
    template_name = 'maestro/create_materia.html'
    success_url = reverse_lazy('control_escolar:maestro_materias')

    login_url = '/login/'
    redirect_field_name = 'redirect_to'

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super(MateriaUpdate, self).get_form_kwargs()
        kwargs.update({'user': self.request.user})
        return kwargs

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'materia'
        context['lista'] = Materia.objects.filter(materia_aula__aula_pertenece=self.request.user)

        return context


class MateriaDelete(LoginRequiredMixin, DeleteView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Materia
    template_name = 'eliminaciones.html'
    success_url = reverse_lazy('control_escolar:maestro_materias')

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'materia'
        context['list_aula'] = Aula.objects.filter(aula_pertenece=self.request.user)

        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected

        return context

class MateriaListaAlumno(TemplateView):
    template_name='maestro/lista_alumno.html'

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'materia'
        context['id_mteria'] = self.kwargs.get('pk')
        mat=Materia.objects.get(materia_id=self.kwargs.get('pk'))
        context['lista_alumno'] = mat.materia_registro_alumnnos.all()

        return context

class DownloadExcel(TemplateView):
    template_name='maestro/lista_alumno.html'

    def get(self, request , *args, **kwargs):
        from openpyxl.utils import get_column_letter
        import calendar
        wb = Workbook()
        ws=wb.active
        # id_pedido=self.kwargs.get('pk')
        # ped = Pedido.objects.get(ped_id_ped=id_pedido)

        mes_get = self.request.GET.get('mes')
        id_materia = self.request.GET.get('id_materia')
        m=mes_get.split('-')
        año=m[0]
        mes=m[1]
        cal= calendar.Calendar()
        materia=Materia.objects.get(materia_id=id_materia)



        ws['A1'] = 'Lista de Materia '+materia.materia_nombre
        st=ws['A1']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')

        st=ws['A2']
        st.font = Font(size=14, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        # ws.merge_cells('A1:F1')

        ws.sheet_properties.tabColor = "1072BA"
        meses={1:'Enero',2:'Febrero',3:'Marzo', 4:'Abril', 5:'Mayo', 6:'Junio', 7:'Julio', 8:'Agosto', 9:'Septiembre', 10:'Octubre', 11:'Noviembre', 12:'Diciembre'}

        ws['A2'] = 'Mes de '+meses[int(mes)]


        ws['A4'] = 'Nombre'
        fin_x=0
        for x in cal.itermonthdays(int(año), int(mes)):
            if x != 0:
                ws.cell(row=4, column=x+1).value = x
                fin_x = x

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=fin_x)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=fin_x)
        lista_alumno=materia.materia_registro_alumnnos.all()
        contador=5
        for alumno in lista_alumno:
            ws.cell(row=contador, column=1).value = alumno.first_name+' '+alumno.last_name
            contador += 1

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+1


        nombre_archivo='Lista_materia.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response


class DocumentoCreate(LoginRequiredMixin, CreateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Documento
    form_class = DocumentoCreateForm
    template_name = 'maestro/docuemnto_crea.html'
    success_url = reverse_lazy('control_escolar:maestro_documento')

    def form_valid(self, form):
        form.instance.doc_pertenece = self.request.user
        return super(DocumentoCreate, self).form_valid(form)

    def get_form_kwargs(self):
        kwargs = super(DocumentoCreate, self).get_form_kwargs()
        kwargs.update({'user': self.request.user})
        return kwargs

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'materia'
        context['list'] = Documento.objects.filter(doc_pertenece=self.request.user)
        return context

class DocumentoUpdate(LoginRequiredMixin, UpdateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Documento
    form_class = DocumentoCreateForm
    template_name = 'maestro/docuemnto_crea.html'
    success_url = reverse_lazy('control_escolar:maestro_documento')

    def get_form_kwargs(self):
        kwargs = super(DocumentoUpdate, self).get_form_kwargs()
        kwargs.update({'user': self.request.user})
        return kwargs

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'materia'
        context['list'] = Documento.objects.filter(doc_pertenece=self.request.user)
        return context
class DocumentoDelete(LoginRequiredMixin, DeleteView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Documento
    template_name = 'eliminaciones.html'
    success_url = reverse_lazy('control_escolar:maestro_documento')

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'materia'


        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected

        return context



class AjaxableResponseMixin:
    """
    Mixin to add AJAX support to a form.
    Must be used with an object-based FormView (e.g. CreateView)
    """
    def form_invalid(self, form):
        response = super().form_invalid(form)
        if self.request.is_ajax():
            return JsonResponse(form.errors, status=400)
        else:
            return response

    def form_valid(self, form):
        # We make sure to call the parent's form_valid() method because
        # it might do some processing (in the case of CreateView, it will
        # call form.save() for example).
        response = super().form_valid(form)
        contenido=''
        if self.request.is_ajax():
            print(form.instance.unidad_materia)

            unidades = Unidad.objects.filter(unidad_materia=form.instance.unidad_materia.materia_id)
            for unidad in unidades:
                contenido += '<tr>' \
                '<td>'+unidad.unidad_nombre +'</td>' \
                '<td><button type="button" onclick="update_unidad('+str(unidad.unidad_id)+')" class="btn btn btn-info"><span class="fas fa-pen-square"></span></button>' \
                '<button type="button" onclick="eliminar_unidad('+str(unidad.unidad_id)+')" class="btn btn-danger"><span class="fas fa-trash"></span></button>' \
                '</tr>'

            return JsonResponse(
                {
                    'content':{
                        'cont': contenido,
                    }
                }
            )
        else:
            return response

class UnidadCreate(LoginRequiredMixin, AjaxableResponseMixin, CreateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Unidad
    form_class = UnidadForm
    template_name = 'maestro/unidad_create.html' 
    success_url = reverse_lazy('control_escolar:maestro_documento')

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)


    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'unidad'
        context['list_aula'] = Aula.objects.filter(aula_pertenece=self.request.user)

        # context['lista'] = Materia.objects.filter(materia_aula__aula_pertenece=self.request.user)
        return context


class Consultas_json(CreateView):
    model = Unidad
    form_class = UnidadForm
    template_name = 'maestro/unidad_create.html'
    success_url = reverse_lazy('control_escolar:maestro_documento')

    def post(self, request, *args, **kwargs):
        tipo_post = request.POST.get('tipo_post')
        contenido='';
        tipo_msn=''
        if tipo_post == 'aula':
            id_aula = request.POST.get('id_aula')
            tipo_msn='aula'
            materias=Materia.objects.filter(materia_aula=id_aula)
            contenido='<option selected disabled>Elija una opcion</option>'
            for materia in materias:
                contenido += '<option value="'+str(materia.materia_id)+'">'+materia.materia_nombre+'</option>';

        elif tipo_post == 'materia':
            id_materia = request.POST.get('id_materia')
            tipo_msn='materia'
            unidades = Unidad.objects.filter(unidad_materia=id_materia)
            for unidad in unidades:
                contenido += '<tr>' \
                '<td>'+unidad.unidad_nombre +'</td>' \
                '<td><button type="button" onclick="update_unidad('+str(unidad.unidad_id)+')" class="btn btn btn-info"><span class="fas fa-pen-square"></span></button>' \
                '<button type="button" onclick="eliminar_unidad('+str(unidad.unidad_id)+')" class="btn btn-danger"><span class="fas fa-trash"></span></button>' \
                ' </td>' \
                '</tr>'

        json = JsonResponse(
            {
                'content':{
                    'contenido': contenido,
                    'tipo_mensaje': tipo_msn,
                }
            }
        )

        return json

class UnidadDelete(TemplateView):

    def post(self, request, *args, **kwargs):
        id_unidad = request.POST.get('id_unidad')
        id_materia = request.POST.get('id_materia')
        unidad=Unidad.objects.get(unidad_id=id_unidad)
        unidad.delete()
        contenido =''
        unidades = Unidad.objects.filter(unidad_materia=id_materia)
        for unidad in unidades:
            contenido += '<tr>' \
            '<td>'+unidad.unidad_nombre +'</td>' \
            '<td><button type="button" onclick="update_unidad('+str(unidad.unidad_id)+')" class="btn btn btn-info"><span class="fas fa-pen-square"></span></button>' \
            '<button type="button" onclick="eliminar_unidad('+str(unidad.unidad_id)+')" class="btn btn-danger"><span class="fas fa-trash"></span></button>' \
            ' </td>' \
            '</tr>'

        json = JsonResponse(
            {
                'content':{
                    'contenido': contenido,
                    'tipo_mensaje': 'materia',
                }
            }
        )

        return json

class UnidadDeleteView(DeleteView):
    model=Unidad
    template_name = 'eliminacionesjs.html'
    success_url = reverse_lazy('control_escolar:maestro_tarea_crear')

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'tarea'


        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected

        return context
        
class UnidadUpdate(TemplateView):

    def post(self, request, *args, **kwargs):
        id_unidad = request.POST.get('id_unidad')
        id_materia = request.POST.get('id_materia')
        cambio_nombre = request.POST.get('cambio_nombre')
        unidad=Unidad.objects.filter(unidad_id=id_unidad).update(unidad_nombre=cambio_nombre)
        contenido =''
        unidades = Unidad.objects.filter(unidad_materia=id_materia)
        for unidad in unidades:
            contenido += '<tr>' \
            '<td>'+unidad.unidad_nombre +'</td>' \
            '<td><button type="button" onclick="update_unidad('+str(unidad.unidad_id)+')" class="btn btn btn-info"><span class="fas fa-pen-square"></span></button>' \
            '<button type="button" onclick="eliminar_unidad('+str(unidad.unidad_id)+')" class="btn btn-danger"><span class="fas fa-trash"></span></button>' \
            ' </td>' \
            '</tr>'

        json = JsonResponse(
            {
                'content':{
                    'contenido': contenido,
                    'tipo_mensaje': 'materia',
                }
            }
        )

        return json





class AjaxableResponseMixinTarea:
    """
    Mixin to add AJAX support to a form.
    Must be used with an object-based FormView (e.g. CreateView)
    """
    def form_invalid(self, form):
        response = super().form_invalid(form)
        if self.request.is_ajax():
            return JsonResponse(form.errors, status=400)
        else:
            return response

    def form_valid(self, form):
        from django.utils.formats import localize
        
        contenido=''
        bandera=False
        if self.request.is_ajax():
            tareas = Tarea.objects.filter(tarea_unidad=form.instance.tarea_unidad.unidad_id)
            get_sum_porcentaje_tarea=Tarea.objects.filter(tarea_unidad=form.instance.tarea_unidad.unidad_id).aggregate(total_porcentaje=Sum('tarea_porcentaje'))
            get_sum_porcentaje_tarea=0 if get_sum_porcentaje_tarea['total_porcentaje'] == None else get_sum_porcentaje_tarea['total_porcentaje']
            porcentaje_input_form=form.instance.tarea_porcentaje
            total = get_sum_porcentaje_tarea + porcentaje_input_form
            if total <= 100:
                response = super().form_valid(form)
                bandera=True
                for tarea in tareas:
                    contenido += '<tr>' \
                    '<td>'+tarea.tarea_nombre +'</td>' \
                    '<td>'+tarea.tarea_descripcion +'</td>' \
                    '<td>'+localize(tarea.tarea_fecha_inicio) +'</td>' \
                    '<td>'+localize(tarea.tarea_fecha_termino) +'</td>' \
                    '<td>'+tarea.get_tarea_tipo_display() +'</td>' \
                    '<td>'+str(tarea.tarea_porcentaje) +'%</td>' \
                    '<td>' \
                    '<a class="btn btn-info" onclick="get_update_tarea('+str(tarea.tarea_id)+')" role="button"><span class="fas fa-pen-square"></span></a>' \
                    '<a class="btn btn-danger" onclick="delete_tarea('+str(tarea.tarea_id)+')" href="#" role="button"><span class="fas fa-trash"></span></a>' \
                    '</tr>'
                tipo_set=form.instance.tarea_tipo
                if tipo_set == 'PARA CALIFICAR':
                    unida=Unidad.objects.get(unidad_id=form.instance.tarea_unidad.unidad_id)
                    idmateria=unida.unidad_materia.materia_id
                    materia=Materia.objects.get(materia_id=idmateria)
                    tarea = form.save(commit=False)
                    tarea.save()
                    for id_user in materia.materia_registro_alumnnos.all():
                        doc=TareaDocumento(tareaDocumento_archivo='s/n',tareaDocumento_comentario_alumno='s/n' ,tareaDocumento_pertenece=id_user ,tareaDocumento_Tarea=tarea)
                        doc.save()
            else:
                bandera=False


            


            return JsonResponse(
                {
                    'content':{
                        'cont': contenido,
                        'acepted': bandera,
                    }
                }
                )
        else:
            return response 




class TareaCreate(LoginRequiredMixin, AjaxableResponseMixinTarea, CreateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Tarea 
    form_class = TareaForm 
    template_name = 'maestro/tarea_crear.html' 
    success_url = reverse_lazy('control_escolar:maestro_tarea_crear')

    def get_form_kwargs(self):
        kwargs = super(TareaCreate, self).get_form_kwargs()
        kwargs.update({'user': self.request.user})
        return kwargs

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs) 


    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'tarea'
        context['list_materia'] = Materia.objects.filter(materia_aula__aula_pertenece=self.request.user)

        # context['lista'] = Materia.objects.filter(materia_aula__aula_pertenece=self.request.user)
        return context

class JsonTareas(TemplateView):
    def post(self, request, *args, **kwargs):
        from django.utils.formats import localize
        tipo_post = request.POST.get('tipo_post')
        contenido='';
        tipo_msn=''
        if tipo_post == 'materia':
            tipo_msn='materia'
            id_materia = request.POST.get('id_materia')
            unidades = Unidad.objects.filter(unidad_materia=id_materia)
            contenido='<option selected disabled>Elija una opcion</option>'
            for unidad in unidades:
                contenido += '<option value="'+str(unidad.unidad_id)+'">'+unidad.unidad_nombre+'</option>';
        elif tipo_post == 'tareas':
            UnidadID = request.POST.get('UnidadID')
            tareas = Tarea.objects.filter(tarea_unidad=UnidadID)
            for tarea in tareas:
                tipo_msn='tarea'
                contenido += '<tr>' \
                '<td>'+tarea.tarea_nombre +'</td>' \
                '<td>'+tarea.tarea_descripcion +'</td>' \
                '<td>'+str(localize(tarea.tarea_fecha_inicio)) +'</td>' \
                '<td>'+str(localize(tarea.tarea_fecha_termino)) +'</td>' \
                '<td>'+tarea.get_tarea_tipo_display() +'</td>' \
                '<td>'+str(tarea.tarea_porcentaje) +'%</td>' \
                '<td>' \
                '<a class="btn btn-info" onclick="get_update_tarea('+str(tarea.tarea_id)+')" role="button"><span class="fas fa-pen-square"></span></a>' \
                '<a class="btn btn-danger" onclick="delete_tarea('+str(tarea.tarea_id)+')" href="#" role="button"><span class="fas fa-trash"></span></a>' \
                '</tr>'




        json = JsonResponse(
            {
                'content':{
                    'contenido': contenido,
                    'tipo_mensaje': tipo_msn,
                }
            }
        )

        return json

class TareaDelete(LoginRequiredMixin, DeleteView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Tarea
    template_name = 'eliminacionesjs.html'
    success_url = reverse_lazy('control_escolar:maestro_tarea_crear')

    

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'tarea'


        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected

        return context








""" Mixin para la actualizacion de tareas"""

class AjaxMixinTareaUpdate:
    """
    Mixin to add AJAX support to a form.
    Must be used with an object-based FormView (e.g. CreateView)
    """
    def form_invalid(self, form):
        response = super().form_invalid(form)
        if self.request.is_ajax():
            return JsonResponse(form.errors, status=400)
        else:
            return response

    def form_valid(self, form):
        # We make sure to call the parent's form_valid() method because
        # it might do some processing (in the case of CreateView, it will
        # call form.save() for example).
        estado = False
        if self.request.is_ajax():
            get_tarea_id=Tarea.objects.get(tarea_id=self.kwargs.get('pk'))
            get_sum_porcentaje_tarea=Tarea.objects.filter(tarea_unidad=form.instance.tarea_unidad.unidad_id).aggregate(total_porcentaje=Sum('tarea_porcentaje'))
            get_sum_porcentaje_tarea=0 if get_sum_porcentaje_tarea['total_porcentaje'] == None else get_sum_porcentaje_tarea['total_porcentaje']
            porcentaje_input_form=form.instance.tarea_porcentaje
            total = (get_sum_porcentaje_tarea - get_tarea_id.tarea_porcentaje) + porcentaje_input_form
            if total <= 100:
                estado = True
                response = super().form_valid(form)
                tipo_set=form.instance.tarea_tipo
                if tipo_set=='PARA CALIFICAR' and get_tarea_id.tarea_tipo == 'ENTREGA':
                    unida=Unidad.objects.get(unidad_id=form.instance.tarea_unidad.unidad_id)
                    idmateria=unida.unidad_materia.materia_id
                    materia=Materia.objects.get(materia_id=idmateria)
                    tarea = form.save(commit=False)
                    tarea.save()
                    for id_user in materia.materia_registro_alumnnos.all():
                        doc=TareaDocumento(tareaDocumento_archivo='s/n',tareaDocumento_comentario_alumno='s/n' ,tareaDocumento_pertenece=id_user ,tareaDocumento_Tarea=tarea)
                        doc.save()

            
            tareas=Tarea.objects.filter(tarea_unidad=self.kwargs.get('id_unidad')).values('tarea_id','tarea_nombre','tarea_descripcion','tarea_fecha_inicio','tarea_fecha_termino','tarea_tipo','tarea_unidad', 'tarea_porcentaje')
            tarea_list=list(tareas)
            data = {
                'estado': estado,
                'Tareas': tarea_list,
            }
            return JsonResponse(data)
        else:
            return response

class TareaUpdateChangue(LoginRequiredMixin,AjaxMixinTareaUpdate, UpdateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Tarea
    form_class = TareaFormEdit
    template_name = 'maestro/tarea_update.html' 
    success_url = reverse_lazy('control_escolar:maestro_tarea_crear')

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)


class TareaUpdate(TemplateView):    
    template_name = 'maestro/tarea_update.html'  

    def get(self, request, *args, **kwargs):
        tarea=Tarea.objects.filter(tarea_id=kwargs.get('pk')).values('tarea_nombre','tarea_descripcion','tarea_fecha_inicio','tarea_fecha_termino','tarea_tipo','tarea_unidad', 'tarea_porcentaje')
        tarea_list=list(tarea)
        return JsonResponse(tarea_list, safe=False)



  
    

class BlogCreate(LoginRequiredMixin, CreateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Blog
    form_class = BlogFrom
    template_name = 'maestro/blog_create.html'
    success_url = reverse_lazy('control_escolar:maestro_blog_list')

    def form_valid(self, form):
        instancia = form.save(commit=False)
        instancia.blog_pertenece = self.request.user
        instancia.save()
        return super(BlogCreate, self).form_valid(form)

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super(BlogCreate, self).get_form_kwargs()
        kwargs.update({'user': self.request.user})
        return kwargs

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'blog'
        return context

class BolgList(LoginRequiredMixin, ListView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Blog
    paginate_by = 8
    template_name = 'maestro/blog_list.html'



    def get_context_data(self, **kwargs): 
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'blog'
        return context

    def get_queryset(self, *args, **kwargs):
        queryset = super(BolgList, self).get_queryset()
        queryset = queryset.filter(blog_pertenece=self.request.user)
        return queryset

class BlogUpdate(LoginRequiredMixin, UpdateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Blog
    form_class = BlogFrom
    template_name = 'maestro/blog_create.html'
    success_url = reverse_lazy('control_escolar:maestro_blog_list')


    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super(BlogUpdate, self).get_form_kwargs()
        kwargs.update({'user': self.request.user})
        return kwargs


    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'blog'
        return context

class BlogDelete(LoginRequiredMixin, DeleteView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Blog
    template_name = 'eliminaciones.html'
    success_url = reverse_lazy('control_escolar:maestro_blog_list')

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'blog'


        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected

        return context

class BlogDetalle(LoginRequiredMixin, DetailView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Blog
    template_name = 'maestro/blog_detalle.html'


    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'blog'
        context['coment_form'] = ComentarioBlogForm()
        context['id_blog'] = self.kwargs['pk']
        context['comentarios'] = ComentarioBlog.objects.filter(comentario_blog=self.kwargs['pk'])

        return context

class ComentarioCreate(LoginRequiredMixin, CreateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = ComentarioBlog
    form_class = ComentarioBlogForm
    template_name = 'maestro/blog_detalle.html'
    success_url = reverse_lazy('control_escolar:maestro_blog_list')

    def form_valid(self, form):
        instancia = form.save(commit=False)
        instancia.comentario_comentado_by = self.request.user
        blog=Blog.objects.get(blog_id=int(self.kwargs['pk']))
        instancia.comentario_blog = blog
        instancia.save()
        return redirect('/maestro/blog/detalle/'+str(self.kwargs['pk']))


class SelectAula(TemplateView):
    template_name = 'maestro/select_aula.html'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'calificar'
        context['aula'] = Aula.objects.filter(aula_pertenece=self.request.user)
        context['colors'] = ['primary', 'secondary', 'success', 'danger', 'warning', 'info', 'light', 'dark']
        
        return context
class SelectMateria(TemplateView):
    template_name = 'maestro/select_materia.html'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'calificar'
        context['materias'] = Materia.objects.filter(materia_aula=self.kwargs.get('aula'))
        context['colors'] = ['primary', 'secondary', 'success', 'danger', 'warning', 'info', 'light', 'dark']
        
        return context
class SelectUnidad(TemplateView):
    template_name = 'maestro/select_unidad.html'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'calificar'
        context['unidades'] = Unidad.objects.filter(unidad_materia=self.kwargs.get('materia'))
        context['colors'] = ['primary', 'secondary', 'success', 'danger', 'warning', 'info', 'light', 'dark']
        
        return context
class SelectTarea(TemplateView):
    template_name = 'maestro/select_tarea.html'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'calificar'
        context['Tareas'] = Tarea.objects.filter(tarea_unidad=self.kwargs.get('unidad'))
        context['colors'] = ['primary', 'secondary', 'success', 'danger', 'warning', 'info', 'light', 'dark']
        
        return context

class TareaEntregadaList(LoginRequiredMixin, ListView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = TareaDocumento
    template_name = 'maestro/calificar_tarea.html' 

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_queryset(self):
        queryset = super(TareaEntregadaList, self).get_queryset()
        queryset=queryset.filter(tareaDocumento_Tarea=self.kwargs.get('id_tarea'))
        cambios = False
        status=self.request.GET.get('estado')
        if status == '2' or status == None:
            pass
        else:
            queryset=queryset.filter(tareaDocumento_status=status)

        # if self.request.method == 'GET':
        #     filter = self.request.GET.get('estado')
        #     materia = self.request.GET.get('materia')
        #     if filter == '0':
        #         cambios = True
        #         queryset=queryset.filter(tareaDocumento_status=False)
        #     if filter == '1':
        #         cambios = True
        #         queryset=queryset.filter(tareaDocumento_status=True)
        #     if filter == '2':
        #         cambios = True
        #         pass
        #     if materia != None:
        #         cambios = True
        #         queryset=queryset.filter(tareaDocumento_Tarea__tarea_unidad__unidad_materia__materia_id=materia)
        #     if cambios == False:
        #         queryset=queryset.filter(tareaDocumento_Tarea__tarea_unidad__unidad_materia__materia_aula__aula_pertenece=self.request.user, tareaDocumento_status=False)


        return queryset

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'calificar'
        context['get_tarea'] = Tarea.objects.get(tarea_id=self.kwargs.get('id_tarea'))

        return context



class TareaEntregadaUpdate(LoginRequiredMixin, UpdateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = TareaDocumento
    form_class = TareaEntregadaEdit
    template_name = 'maestro/blog_create.html'
    success_url = reverse_lazy('control_escolar:maestro_tarea_entregada_list')

    def get_success_url(self):
        url_new=reverse_lazy('control_escolar:maestro_tarea_entregada_list', kwargs={'id_tarea': self.kwargs.get('id_tarea')},)
        print(url_new)
        url_get=self.request.GET.urlencode()
        return url_new+'?'+url_get

    def form_valid(self, form):
        form.instance.tareaDocumento_status = True
        return super(TareaEntregadaUpdate, self).form_valid(form)

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'calificar'

        return context
class MateriaGet(View):
    def get(self, request, *args, **kwargs):
        aula=request.GET.get('aula')
        materia=request.GET.get('materia')

        materias=Materia.objects.filter(materia_aula=aula).values('materia_id','materia_nombre')
        unidades=Unidad.objects.filter(unidad_materia=materia).values()

        materia_list=list(materias)
        unidad_list=list(unidades)
        
        data={
            'materia':materia_list,
            'unidades':unidad_list
        }
        return JsonResponse(data)

class PromediarUnidad(View):
    def get(self, request, *args, **kwargs):
        aula = self.request.GET.get('aula')
        materia = self.request.GET.get('materia')
        unidad = self.request.GET.get('unidad')
        
        if unidad != None:
            unidad_q=Unidad.objects.get(unidad_id=unidad)
            materia_q=Materia.objects.get(materia_id=unidad_q.unidad_materia.materia_id)
            # context['alumnos'] = materia_q.materia_registro_alumnnos.all()
            total_tareas = Tarea.objects.filter(tarea_unidad=unidad).count()
            for al in materia_q.materia_registro_alumnnos.all():
                entrego = TareaDocumento.objects.filter(tareaDocumento_Tarea__tarea_unidad=unidad, tareaDocumento_pertenece=al).count()
                falta_califica = TareaDocumento.objects.filter(tareaDocumento_Tarea__tarea_unidad=unidad, tareaDocumento_pertenece=al, tareaDocumento_status = False).count()
                Suma_tareas = TareaDocumento.objects.filter(tareaDocumento_Tarea__tarea_unidad=unidad, tareaDocumento_pertenece=al).aggregate(suma_total=Sum((F('tareaDocumento_Tarea__tarea_porcentaje') * F('tareaDocumento_calificacion'))/100, output_field=FloatField()))
                if Suma_tareas['suma_total'] == None:
                    Suma_tareas['suma_total'] = 0
                calificacion =round(Suma_tareas['suma_total'], 2)
                CalificacionUnidad.objects.filter(calU_unidadID=unidad_q.unidad_id,calU_materiaID=materia_q.materia_id,calU_pertenece=al).delete()
                calficar = CalificacionUnidad(
                    calU_Materia_nombre = materia_q.materia_nombre,
                    calU_unidadNombre=unidad_q.unidad_nombre,
                    calU_unidadID=unidad_q.unidad_id,
                    calU_materiaID=materia_q.materia_id,
                    calU_calificacion=calificacion,
                    calU_entrego=str(entrego)+' de '+str(total_tareas) ,
                    calU_pertenece=al,
                    calU_falta_calificar=falta_califica,
                )
                calficar.save()
            calificaciones = CalificacionUnidad.objects.filter(calU_unidadID=unidad).values(
                'calU_Materia_nombre',
                'calU_unidadNombre',
                'calU_calificacion', 
                'calU_entrego',
                'calU_falta_calificar',
                'calU_pertenece__first_name',
                'calU_pertenece__last_name',
                )
            cali_list=list(calificaciones)
        data={
            'calificaciones':cali_list,
        }
        return JsonResponse(data)
 
class PromedioUnidadDetalle_xls(View):
    def get(self, request , *args, **kwargs):
        from openpyxl.utils import get_column_letter
        import calendar
        import random
        from openpyxl.styles import PatternFill
        from openpyxl.styles.borders import Border, Side 
        
        wb = Workbook()
        ws=wb.active
        # id_pedido=self.kwargs.get('pk')
        # ped = Pedido.objects.get(ped_id_ped=id_pedido)

        unidad_id=self.request.GET.get('unidad_id')
        unidad_get=Unidad.objects.get(unidad_id=unidad_id)
        tareas_count=Tarea.objects.filter(tarea_unidad=unidad_id).count()

        ws['A1'] = "{} -- {}".format(unidad_get.unidad_materia.materia_nombre, unidad_get.unidad_nombre)
        st=ws['A1']
        st.font = Font(size=12, b=True, color="004ee0")
        st.alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        
        #COLUMNAS EXCEL
        ws['A2']='Alumno'
        ws['B2']='Tarea'
        ws['C2']='Calificacion'
        ws['D2']='%'
        doc_alum=TareaDocumento.objects.filter(tareaDocumento_Tarea__tarea_unidad=unidad_get).order_by('tareaDocumento_pertenece')
        colors_rbga=['EFA8D3', 'BEABF6', 'F3B8F8', '6FE3F5', 'F4B2B2', 'FFB7D5', 'A7ABFA', '94F4B5', 'F7D896', 'D4D3D3', 'FFBCBC', 'F8CEA4', 'CEF683', 'C8FBC0', '87C8E2', 'A7C4FA', 'F49495', 'D4B1EB']
        

        thin_border = Border(left=Side(style='thin', color='FFFFFF'), 
        right=Side(style='thin', color='FFFFFF'), 
        top=Side(style='thin', color='FFFFFF'), 
        bottom=Side(style='thin', color='FFFFFF')) 
        

        contador=3
        contador_tarea_reset=1
        doc_count=0
        tem_cont_ped = 0
        for alumno in doc_alum:
            if alumno.tareaDocumento_pertenece.id != tem_cont_ped:
                color_temporal_gen=random.choice(colors_rbga)  
                contador += 1
            tem_cont_ped=alumno.tareaDocumento_pertenece.id

            ws.cell(row=contador, column=1).fill  = PatternFill(start_color=color_temporal_gen, end_color=color_temporal_gen, fill_type = 'solid')
            ws.cell(row=contador, column=2).fill  = PatternFill(start_color=color_temporal_gen, end_color=color_temporal_gen, fill_type = 'solid')
            ws.cell(row=contador, column=3).fill  = PatternFill(start_color=color_temporal_gen, end_color=color_temporal_gen, fill_type = 'solid')
            ws.cell(row=contador, column=4).fill  = PatternFill(start_color=color_temporal_gen, end_color=color_temporal_gen, fill_type = 'solid')

            ws.cell(row=contador, column=1).border  = thin_border
            ws.cell(row=contador, column=2).border  = thin_border
            ws.cell(row=contador, column=3).border  = thin_border
            ws.cell(row=contador, column=4).border  = thin_border
            
            ws.cell(row=contador, column=1).value = alumno.tareaDocumento_pertenece.get_full_name()
            ws.cell(row=contador, column=2).value = alumno.tareaDocumento_Tarea.tarea_nombre
            ws.cell(row=contador, column=3).value = alumno.tareaDocumento_calificacion
            calific_alumno= 0 if alumno.tareaDocumento_calificacion == None else alumno.tareaDocumento_calificacion
            porcentaje=(alumno.tareaDocumento_Tarea.tarea_porcentaje * calific_alumno)/10
            ws.cell(row=contador, column=4).value = "{}% de {}%".format(porcentaje, alumno.tareaDocumento_Tarea.tarea_porcentaje)
            # ws.cell(row=contador, column=4).value = "{}% de {}%".format((alumno.tareaDocumento_Tarea.tarea_porcentaje * alumno.tareaDocumento_calificacion)/10, alumno.tareaDocumento_Tarea.tarea_porcentaje)
            # if contador_tarea_reset == doc_count:
            #     ws.cell(row=contador+1, column=2).value = 'Promedio'
            #     ws.cell(row=contador+1, column=3).value = '=SUM(C{}:C{})/{}'.format(contador-(contador_tarea_reset-1),contador, tareas_count)
            #     contador += 1
            #     contador_tarea_reset = 1
            
            doc_count=TareaDocumento.objects.filter(tareaDocumento_Tarea__tarea_unidad=unidad_get, tareaDocumento_pertenece=alumno.tareaDocumento_pertenece).count()
            contador += 1
            contador_tarea_reset += 1

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    if cell.row != 1:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[get_column_letter(col)].width = value+2


        nombre_archivo='calificaciones.xls'
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition']=content
        wb.save(response)
        return response

class PromediarCreate(LoginRequiredMixin, TemplateView): 
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    template_name = 'maestro/promediar.html'

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'promediar'
        context['aulas'] = Aula.objects.filter(aula_pertenece=self.request.user)
        context['msn'] = 'Elija un Aula'       

        return context
    
class PromediarMateria(LoginRequiredMixin, TemplateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    template_name = 'maestro/promediar_materia.html'

    def dispatch(self, *args, **kwargs):
        if self.request.user.is_authenticated:
            if self.request.user.is_alumno:
                return redirect('/')
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'promediar'
        context['aulas'] = Aula.objects.filter(aula_pertenece=self.request.user)
        context['msn'] = '<ol><h1><li>Elija un aula</li></h1></ol>'

        if self.request.method == 'GET':
            aula = self.request.GET.get('aula')
            materia = self.request.GET.get('materia')
            context['materias'] = Materia.objects.filter(materia_aula=aula)
            if aula != None: 
                context['msn'] = '<ol start="2"><h1><li>Elija una Materia</li></h1></ol>'
            if materia != None:
                # OBTENEMOS EL OBJECTO DE MATERIA
                materia_q=Materia.objects.get(materia_id=materia)
                context['msn'] = '<ol><h1><li>Elija un aula</li></h1></ol>'
                context['materia_txt'] = materia_q.materia_nombre
                total_unidades_materia=Unidad.objects.filter(unidad_materia=materia).count()
                for cal_al in materia_q.materia_registro_alumnnos.all():
                    suma_total_unidades_alumno=CalificacionUnidad.objects.filter(calU_materiaID=materia, calU_pertenece=cal_al).aggregate(suma_total=Sum('calU_calificacion'))
                    if suma_total_unidades_alumno['suma_total'] == None:
                        suma_total_unidades_alumno['suma_total'] = 0
                    calficacion = suma_total_unidades_alumno['suma_total']/total_unidades_materia
                    find_mat_alum = CalificacionMateria.objects.filter(calM_materiaID = materia_q.materia_id, calM_alumno=cal_al).delete()
                    calificacion_final=CalificacionMateria(
                        calM_nombre_materia=materia_q.materia_nombre,
                        calM_materiaID=materia_q.materia_id,
                        calM_calificacionFinal=calficacion,
                        calM_alumno=cal_al,
                    )
                    calificacion_final.save()
            context['alumnos'] = CalificacionMateria.objects.filter(calM_materiaID=materia)
        return context



"""
#########################################################################################################
#########################################################################################################
"""









"""
CLASES ESCUSIVAMENTE PARA LOS USUARIOS CON PERMISOS DE ALUMNO
DE AQUI EN ADELANTE
#########################################################################################################
#########################################################################################################
#########################################################################################################
"""

class InscripcionCreate(LoginRequiredMixin, View):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'

    def post(self, request, *args, **kwargs):
        regis_materia = request.POST.get('regis_materia')
        try:
            materia=Materia.objects.get(materia_id=regis_materia)
            materia.materia_registro_alumnnos.add(self.request.user)
            link_materia=reverse_lazy('control_escolar:alumno_materia_list')

            tareas=Tarea.objects.filter(tarea_unidad__unidad_materia=regis_materia,tarea_tipo='PARA CALIFICAR')
            for tarea in tareas:
                doc=TareaDocumento(tareaDocumento_archivo='s/n',tareaDocumento_comentario_alumno='s/n' ,tareaDocumento_pertenece=self.request.user ,tareaDocumento_Tarea=tarea)
                doc.save()
            
        except IntegrityError as err:
            msn="Usted ya se encuentra registrado actualmente en la materia"
            return render_to_response("IntegrityError.html", {"message": msn, 'alumno':self.request.user.is_alumno, 'maestro':self.request.user.is_maestro, 'foto_perfil':self.request.user.foto_perfil})

        return redirect(link_materia)



    # def form_valid(self, form):
    #     instancia = form.save(commit=False)
    #     try:
    #         registro=instancia.regis_materia
    #         materia=Materia.objects.get(materia_id=registro)
    #         materia.materia_registro_alumnnos.add(self.request.user)

    #         return super(InscripcionCreate, self).form_valid(form)
    #     except IntegrityError as e:
    #         msn="Usted ya se encuentra registrado actualmente en la materia"
    #         return render_to_response("IntegrityError.html", {"message": msn})

    # def form_invalid(self, form):
    #     response = super().form_invalid(form)
    #     return redirect('alumno/materias/')

class MateriaAlumnoList(LoginRequiredMixin, ListView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Materia
    template_name = 'alumno/materias_list.html'

    def get_queryset(self, *args, **kwargs):
        queryset = super(MateriaAlumnoList, self).get_queryset()
        queryset = queryset.filter(materia_registro_alumnnos=self.request.user)
        return queryset

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'materia'

        return context

class DocumentMateriaAlumnoList(LoginRequiredMixin, ListView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Materia
    template_name = 'alumno/dcu_materia.html'

    def get_queryset(self, *args, **kwargs):
        queryset = super(DocumentMateriaAlumnoList, self).get_queryset()
        mat_terID = self.kwargs['pk']
        queryset = queryset.filter(materia_id=mat_terID)
        return queryset

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'materia'
        context['color_list'] = ['list-group-item-primary', 'list-group-item-secondary', 'list-group-item-success', 'list-group-item-danger', 'list-group-item-warning', 'list-group-item-info', 'list-group-item-light', 'list-group-item-dark']

        return context

class BoligAlumnoList(LoginRequiredMixin, ListView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = Blog
    template_name = 'alumno/blog_al_list.html'
    paginate_by = 10

    def get_queryset(self, *args, **kwargs):
        queryset = super(BoligAlumnoList, self).get_queryset()

        queryset = queryset.filter(blog_materia__materia_registro_alumnnos=self.request.user)
        # queryset = queryset.filter()
        return queryset

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'blog'

        return context




class AjaxableResponseMixinTareaAlumno:
    """
    Mixin to add AJAX support to a form.
    Must be used with an object-based FormView (e.g. CreateView)
    """
    def form_invalid(self, form):
        response = super().form_invalid(form)
        if self.request.is_ajax():
            return JsonResponse(form.errors, status=400)
        else:
            return response

class TareaAlumnoCreate(LoginRequiredMixin, AjaxableResponseMixinTareaAlumno, CreateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    model = TareaDocumento
    form_class = TareaDocumentoFrom
    template_name = 'alumno/tarea_create.html'  
    success_url = reverse_lazy('control_escolar:alumno_tarea_create')

    def get_form_kwargs(self):
        kwargs = super(TareaAlumnoCreate, self).get_form_kwargs()
        kwargs.update({'user': self.request.user})
        return kwargs

    def form_valid(self, form):
        instancia = form.save(commit=False)
        tipo = False
        try:
            instancia.tareaDocumento_pertenece = self.request.user
            instancia.save()
            msn="Creeado Correctamente"
            tipo=True
            return JsonResponse(
                {
                    'content':{
                        'cont': msn,
                        'tipo': tipo,
                    }
                }
            )
        except IntegrityError as e:
            msn="Esta Tarea se encuentra entregada"

            return JsonResponse(
                {
                    'content':{
                        'cont': msn,
                    }
                }
            )

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'TareaAlumno'
        context['al_maetria'] = Materia.objects.filter(materia_registro_alumnnos=self.request.user)

        return context

class ListTareaAlumno(LoginRequiredMixin, View):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    def post(self, request, *args, **kwargs):
        from django.utils.formats import localize
        from datetime import datetime
        from django.core.exceptions import ObjectDoesNotExist
        tipo_post = request.POST.get('tipo_post')
        contenido='';
        coment='';
        tipo_msn=''
        fecha_actual = datetime.now().date()

        if tipo_post == 'tareas': 
            UnidadID = request.POST.get('UnidadID')
            tareas = Tarea.objects.filter(tarea_unidad=UnidadID)
            for tarea in tareas:
                temporal=''
                temporal_porcentaje=''
                entregado=False
                try:
                    documento=TareaDocumento.objects.get(tareaDocumento_pertenece=self.request.user,tareaDocumento_Tarea=tarea)
                    temporal='<span class="badge badge-primary">'+str(documento.tareaDocumento_calificacion)+'</span>'
                    temporal_porcentaje="""{}% de {}%""".format(documento.porcentaje_tarea(), documento.tareaDocumento_Tarea.tarea_porcentaje)
                    entregado=True
                    coment=documento.tareaDocumento_comentario_maestro
                    if coment == None:
                        coment=''
                except ObjectDoesNotExist:
                    temporal='<span class="badge badge-warning">No entregado</span>'
                    entregado=False
                    coment=''
                
                tipo_msn='tarea'
                contenido += '<tr>' \
                '<td>'+tarea.tarea_nombre +'</td>' \
                '<td>'+tarea.tarea_descripcion +'</td>' \
                '<td>'+coment+'</td>' \
                '<td>'+str(localize(tarea.tarea_fecha_inicio)) +'</td>' \
                '<td>'+str(localize(tarea.tarea_fecha_termino)) +'</td>' \
                '<td>'+str(temporal)+'</td>' \
                '<td>'+temporal_porcentaje+'</td>' \
                '<td>'+tarea.tarea_tipo+'</td>' \
                '<td>'
                fecha_tarea=datetime.strptime(str(tarea.tarea_fecha_termino), "%Y-%m-%d")
                if fecha_actual <= fecha_tarea.date() :
                    if tarea.tarea_tipo == 'PARA CALIFICAR':
                        contenido +='<span class="badge badge-secondary" title="No es necesario entregar">No es necesario entregar</span>'
                    else:
                        if entregado:
                            contenido +='<span class="badge badge-success">Entregado</span>'
                        else:
                            contenido += '<button type="button" onclick="entregar('+str(tarea.tarea_id)+')"  data-toggle="modal" data-target="#Tarea_modal" class="btn btn-default" aria-label="Left Align"><span class="glyphicon glyphicon-pencil" aria-hidden="true"></span>Entregar</button> </tr>'                          

                else:
                    if entregado:
                        contenido +='<span class="badge badge-success">Entregado</span>'
                    else:
                        contenido += '<span class="badge badge-danger">Tiempo de entrega agotado</span></tr>'
                    






        json = JsonResponse(
            {
                'content':{
                    'contenido': contenido,
                    'tipo_mensaje': tipo_msn,
                }
            }
        )

        return json
class CalificacionesVer(LoginRequiredMixin, TemplateView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'
    template_name = 'alumno/calificaciones.html'

    def get_context_data(self, **kwargs):
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)
        context['maestro'] = self.request.user.is_maestro
        context['alumno'] = self.request.user.is_alumno
        context['foto_perfil'] = self.request.user.foto_perfil
        context['Usuario'] = self.request.user
        context['activate'] = 'calificacion'
        context['unidades'] = CalificacionUnidad.objects.filter(calU_pertenece=self.request.user)
        context['calificaciones'] = CalificacionMateria.objects.filter(calM_alumno=self.request.user)

        return context


class ResponseMaestroAjax(TemplateView):
    template_name='maestro/ma_index.html'
    def get(self, request, *args, **kwargs):
        fin = timezone.now()
        inicio = fin - timedelta(days=2)

        data =  dict()

        if self.request.user.is_maestro:
            coment =  list(ComentarioBlog.objects.filter(comentario_blog__blog_pertenece=self.request.user, comentario_creado__range=[inicio, fin]).values('comentario_comentario', 'comentario_comentado_by__first_name','comentario_comentado_by__last_name', 'comentario_comentado_by__foto_perfil', 'comentario_creado'))
            data['coment'] = coment
            data['number_coment'] = ComentarioBlog.objects.filter(comentario_blog__blog_pertenece=self.request.user, comentario_creado__range=[inicio, fin]).count()
        else:
            coment =  list(ComentarioBlog.objects.filter(comentario_blog__blog_materia__materia_registro_alumnnos=self.request.user, comentario_creado__range=[inicio, fin]).values('comentario_comentario', 'comentario_comentado_by__first_name','comentario_comentado_by__last_name', 'comentario_comentado_by__foto_perfil', 'comentario_creado'))
            data['coment'] = coment
            data['number_coment'] = ComentarioBlog.objects.filter(comentario_blog__blog_materia__materia_registro_alumnnos=self.request.user, comentario_creado__range=[inicio, fin]).count()



        return JsonResponse(data)


"""
#########################################################################################################
#########################################################################################################
#########################################################################################################
"""



"""
Nuevas vistas para la nueva version de sensei
"""


class AdminIndex(TemplateView):
    template_name = "dasboard/base.html"


class AulaList(ListView):
    model = Aula
    template_name = "dasboard/maestro/aula.html"

class AulaCreateView(CreateView):
    model = Aula
    template_name = "dasboard/maestro/aula_crear.html"
    form_class = AulaForm
    success_url = reverse_lazy('ctr:list_aula')

    def form_valid(self, form):
        form.instance.aula_pertenece = self.request.user
        return super(AulaCreate, self).form_valid(form)



class AulaDeleteView(LoginRequiredMixin, DeleteView):
    login_url = '/login/'
    redirect_field_name = 'redirect_to'

    model = Aula
    template_name = 'DeleteForm.html'
    success_url = reverse_lazy('ctr:list_aula')

    # def dispatch(self, *args, **kwargs):
    #     if self.request.user.is_authenticated:
    #         if self.request.user.is_alumno:
    #             return redirect('/')
    #     return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        deletable_objects, model_count, protected = get_deleted_objects([self.object])
        context['deletable_objects']=deletable_objects
        context['model_count']=dict(model_count).items()
        context['protected']=protected
        return context


class AulaUpdateView(UpdateView):
    model = Aula
    template_name = "dasboard/maestro/aula_crear.html"
    form_class = AulaForm
    success_url = reverse_lazy('ctr:list_aula')




class MateriaListView(ListView):
    model = Materia
    template_name = "dasboard/maestro/materia_list.html" 



class MateriaCreateView(CreateView):
    model = Materia
    template_name = "FormCreate.html"
    form_class = MateriaForm
    success_url = reverse_lazy('ctr:materia_create')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.update({'user': self.request.user}) 
        return kwargs
